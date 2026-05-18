#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
process_excel_files.py

Batch-extract metadata + records from many .xlsx files into two flat tables
ready for SQL Server ingestion (XLSX + CSV outputs).

Usage:
    python process_excel_files.py --input "C:\excel_files" --output "C:\excel_output"
"""

import os
import re
import glob
import argparse
import logging
import posixpath
import zipfile
import numbers
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook


# ==============================
# ARGUMENTS
# ==============================
def parse_args():
    parser = argparse.ArgumentParser(description="Process Excel files into structured datasets.")
    parser.add_argument("--input", required=True, help="Input folder containing .xlsx files")
    parser.add_argument("--output", required=True, help="Output folder for results")
    parser.add_argument(
        "--export-sql",
        default="",
        help="Optional folder to write offline SQL Server scripts (schema + inserts).",
    )
    return parser.parse_args()


# ==============================
# CONSTANTS
# ==============================
METADATA_FIELDS = [
    "تاریخ انتشار",
    "شماره صفحه",
    "کد پروژه",
    "نام سازنده",
    "شماره بازنگری",
    "تهیه کننده",
    "تایید کننده",
    "نام قطعه",
    "شماره فنی",
    "نام خودرو",
]

BASE_RECORD_FIELDS = [
    "مشخصات کنترلی",
    "شرح کنترل",
    "درجه اهمیت",
    "استاندارد مرجع",
    "محدوده قابل قبول",
    "روش نمونه گیری (تناوب - تعداد) توضیحات",
]

OPTIONAL_RECORD_FIELDS = [
    "نام ابزار/دقت ابزار",
]

RECORD_FIELDS = BASE_RECORD_FIELDS + OPTIONAL_RECORD_FIELDS

INCOMPLETE_COLUMNS = [
    "FileID",
    "FileName",
    "SheetName",
    "ExcelRow",
    "FieldType",
    "Column",
    "IssueType",
    "CellHint",
    "Details",
]

RECORD_OUTPUT_COLUMNS = ["FileID", "شیت"] + BASE_RECORD_FIELDS + OPTIONAL_RECORD_FIELDS

IMAGE_OUTPUT_COLUMNS = [
    "ImageID",
    "FileID",
    "FileName",
    "شیت",
    "ImageName",
    "ImagePath",
    "EmbeddedPath",
]


# ==============================
# UTILS
# ==============================
def normalize(text):
    if text is None:
        return ""
    return str(text).strip().replace("ي", "ی").replace("ك", "ک")


def extract_text_only(cell_or_value):
    # openpyxl stores images separately from cell.value.
    # We intentionally read only the textual/scalar cell value and ignore images.
    value = getattr(cell_or_value, "value", cell_or_value)
    return normalize(value)


def extract_part_number(text):
    text = extract_text_only(text).upper()
    match = re.search(r"\b[0-9A-Z]{6,}(?:-[0-9A-Z]{2,})+\b", text)
    return match.group(0) if match else ""


def make_issue(sheet_name, excel_row, field_type, column_name, issue_type, cell_hint="", details=""):
    return {
        "SheetName": sheet_name or "",
        "ExcelRow": excel_row if excel_row is not None else "",
        "FieldType": field_type,
        "Column": column_name,
        "IssueType": issue_type,
        "CellHint": cell_hint,
        "Details": details,
    }


def is_dimensional_sheet(sheet_name):
    return "ابعادی" in normalize(sheet_name)


def get_header_aliases(target):
    aliases = {
        "مشخصات کنترلی": ["مشخصات کنترلی", "مشخصات", "کنترلی"],
        "شرح کنترل": ["شرح کنترل", "شرح", "کنترل"],
        "درجه اهمیت": ["درجه اهمیت", "اهمیت"],
        "استاندارد مرجع": ["استاندارد مرجع", "مرجع", "استاندارد"],
        "محدوده قابل قبول": ["محدوده قابل قبول", "قابل قبول", "محدوده"],
        "روش نمونه گیری (تناوب - تعداد) توضیحات": [
            "روش نمونه گیری",
            "نمونه گیری",
            "تناوب",
            "توضیحات",
        ],
        "نام ابزار/دقت ابزار": ["نام ابزار/دقت ابزار", "نام ابزار", "دقت ابزار", "ابزار"],
    }
    return aliases.get(target, [target])


def header_matches(target, col_text):
    return any(alias in col_text for alias in get_header_aliases(target))


def safe_folder_name(file_id):
    return f"file_{file_id:05d}"


def normalize_zip_path(base_dir, target):
    base_dir = posixpath.dirname(base_dir)
    return posixpath.normpath(posixpath.join(base_dir, target))


def parse_xml_bytes(zf, path):
    try:
        return ET.fromstring(zf.read(path))
    except KeyError:
        return None


def workbook_sheet_paths(zf):
    ns_main = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    workbook_root = parse_xml_bytes(zf, "xl/workbook.xml")
    rels_root = parse_xml_bytes(zf, "xl/_rels/workbook.xml.rels")
    if workbook_root is None or rels_root is None:
        return {}

    rel_map = {}
    for rel in rels_root:
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rel_map[rel_id] = normalize_zip_path("xl/workbook.xml", target)

    sheet_map = {}
    for sheet in workbook_root.findall("main:sheets/main:sheet", ns_main):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        sheet_path = rel_map.get(rel_id)
        if name and sheet_path:
            sheet_map[name] = sheet_path

    return sheet_map


def drawing_paths_for_sheet(zf, sheet_path):
    ns_main = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheet_root = parse_xml_bytes(zf, sheet_path)
    if sheet_root is None:
        return []

    drawing_rel_ids = []
    for node in sheet_root.findall(".//main:drawing", ns_main):
        rel_id = node.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if rel_id:
            drawing_rel_ids.append(rel_id)

    if not drawing_rel_ids:
        return []

    sheet_rels_path = posixpath.join(
        posixpath.dirname(sheet_path),
        "_rels",
        posixpath.basename(sheet_path) + ".rels",
    )
    rels_root = parse_xml_bytes(zf, sheet_rels_path)
    if rels_root is None:
        return []

    drawing_paths = []
    for rel in rels_root:
        if rel.attrib.get("Id") in drawing_rel_ids:
            target = rel.attrib.get("Target")
            if target:
                drawing_paths.append(normalize_zip_path(sheet_path, target))

    return drawing_paths


def image_targets_for_drawing(zf, drawing_path):
    xdr_ns = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
    drawing_root = parse_xml_bytes(zf, drawing_path)
    if drawing_root is None:
        return []

    drawing_rels_path = posixpath.join(
        posixpath.dirname(drawing_path),
        "_rels",
        posixpath.basename(drawing_path) + ".rels",
    )
    rels_root = parse_xml_bytes(zf, drawing_rels_path)
    if rels_root is None:
        return []

    rel_map = {}
    for rel in rels_root:
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rel_map[rel_id] = normalize_zip_path(drawing_path, target)

    images = []
    anchors = drawing_root.findall(".//xdr:oneCellAnchor", xdr_ns)
    anchors += drawing_root.findall(".//xdr:twoCellAnchor", xdr_ns)
    anchors += drawing_root.findall(".//xdr:absoluteAnchor", xdr_ns)

    for anchor in anchors:
        from_node = anchor.find("xdr:from", xdr_ns)
        row_num = ""
        col_num = ""
        if from_node is not None:
            row_node = from_node.find("xdr:row", xdr_ns)
            col_node = from_node.find("xdr:col", xdr_ns)
            if row_node is not None and row_node.text is not None:
                row_num = str(int(row_node.text) + 1)
            if col_node is not None and col_node.text is not None:
                col_num = str(int(col_node.text) + 1)

        for blip in anchor.findall(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"):
            rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            media_path = rel_map.get(rel_id, "")
            if media_path:
                images.append(
                    {
                        "embedded_path": media_path,
                        "anchor_row": row_num,
                        "anchor_col": col_num,
                    }
                )

    return images


def extract_images_from_excel(file_path, file_id, output_dir, start_image_id):
    image_rows = []
    image_id = start_image_id
    images_root = os.path.join(output_dir, "extracted_images", safe_folder_name(file_id))
    os.makedirs(images_root, exist_ok=True)

    with zipfile.ZipFile(file_path) as zf:
        sheet_map = workbook_sheet_paths(zf)
        assigned_media = set()

        for sheet_name, sheet_path in sheet_map.items():
            drawing_paths = drawing_paths_for_sheet(zf, sheet_path)
            for drawing_path in drawing_paths:
                for image_info in image_targets_for_drawing(zf, drawing_path):
                    embedded_path = image_info["embedded_path"]
                    if embedded_path in assigned_media:
                        continue
                    assigned_media.add(embedded_path)

                    image_name = os.path.basename(embedded_path)
                    output_path = os.path.join(images_root, f"{image_id}_{image_name}")
                    with open(output_path, "wb") as f:
                        f.write(zf.read(embedded_path))

                    image_rows.append(
                        [
                            image_id,
                            file_id,
                            os.path.basename(file_path),
                            sheet_name,
                            image_name,
                            output_path,
                            embedded_path,
                        ]
                    )
                    image_id += 1

        for member in zf.namelist():
            if not member.startswith("xl/media/") or member in assigned_media:
                continue
            image_name = os.path.basename(member)
            output_path = os.path.join(images_root, f"{image_id}_{image_name}")
            with open(output_path, "wb") as f:
                f.write(zf.read(member))

            image_rows.append(
                [
                    image_id,
                    file_id,
                    os.path.basename(file_path),
                    "Unknown",
                    image_name,
                    output_path,
                    member,
                ]
            )
            image_id += 1

    return image_rows, image_id


# ==============================
# METADATA EXTRACTION
# ==============================
def extract_metadata(wb):
    result = {k: "" for k in METADATA_FIELDS}
    issues = []
    key_locations = {k: None for k in METADATA_FIELDS}

    patterns = {
        "تاریخ انتشار": [("تاریخ", "انتشار"), ("انتشار",), ("تاریخ",)],
        "شماره صفحه": [("شماره", "صفحه")],
        "کد پروژه": [("کد", "پروژه")],
        "نام سازنده": [("نام", "سازنده")],
        "شماره بازنگری": [("شماره", "بازنگری")],
        "تهیه کننده": [("تهیه", "کننده")],
        "تایید کننده": [("تایید", "کننده")],
        "نام قطعه": [("نام", "قطعه")],
        "شماره فنی": [("شماره", "فنی")],
        "نام خودرو": [("نام", "خودرو")],
    }

    for sheet in wb.worksheets[2:]:
        for row in sheet.iter_rows(max_row=200, max_col=50):
            for cell in row:
                text = extract_text_only(cell)

                for field, pats in patterns.items():
                    if result[field]:
                        continue

                    for pat in pats:
                        if all(p in text for p in pat):
                            if not key_locations[field]:
                                key_locations[field] = {
                                    "sheet": sheet.title,
                                    "row": cell.row,
                                    "cell": cell.coordinate,
                                }
                            val = ""

                            # same cell
                            if ":" in text:
                                val = text.split(":")[-1].strip()

                            # next cell
                            if not val:
                                try:
                                    val = extract_text_only(row[cell.column])
                                except Exception:
                                    val = ""

                            if val:
                                if field == "شماره فنی":
                                    val = extract_part_number(val)

                                result[field] = val
                                break

    for field in METADATA_FIELDS:
        if result[field]:
            continue
        location = key_locations.get(field)
        if location:
            issues.append(
                make_issue(
                    location["sheet"],
                    location["row"],
                    "metadata",
                    field,
                    "metadata_value_missing",
                    location["cell"],
                    "Metadata label was found but no extractable value was captured.",
                )
            )
        else:
            issues.append(
                make_issue(
                    "Not found",
                    "",
                    "metadata",
                    field,
                    "metadata_key_not_found",
                    "",
                    "Metadata label was not found in scanned sheets.",
                )
            )

    return result, issues


# ==============================
# RECORDS EXTRACTION
# ==============================
def detect_header(row):
    row_text = [extract_text_only(x) for x in row if x]
    score = sum(1 for col in BASE_RECORD_FIELDS if any(header_matches(col, cell) for cell in row_text))
    return score >= 3


def extract_records(wb, file_id):
    records = []
    issues = []
    empty_stop_threshold = 40

    for sheet in wb.worksheets[2:]:
        target_fields = list(BASE_RECORD_FIELDS)
        if is_dimensional_sheet(sheet.title):
            target_fields += OPTIONAL_RECORD_FIELDS

        header = None
        header_idx = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if detect_header(row):
                header = row
                header_idx = i
                break

        if not header:
            issues.append(
                make_issue(
                    sheet.title,
                    "",
                    "records",
                    "*",
                    "header_not_found",
                    "",
                    "No recognizable record header row was detected in this sheet.",
                )
            )
            continue

        col_map = {}
        for idx, col in enumerate(header):
            col_text = extract_text_only(col)
            for target in target_fields:
                if header_matches(target, col_text):
                    col_map[target] = idx

        for target in target_fields:
            if target not in col_map:
                issues.append(
                    make_issue(
                        sheet.title,
                        header_idx + 1,
                        "records",
                        target,
                        "header_column_missing",
                        "",
                        "Required record column was not found in the detected header row.",
                    )
                )

        empty_streak = 0
        for i, row in enumerate(sheet.iter_rows(min_row=header_idx + 2, values_only=True), start=header_idx + 2):
            values = []

            for col in target_fields:
                idx = col_map.get(col)
                val = extract_text_only(row[idx]) if idx is not None and idx < len(row) else ""
                values.append(val)

            populated = sum(1 for v in values[:len(BASE_RECORD_FIELDS)] if v)
            if populated == 0:
                empty_streak += 1
                if empty_streak >= empty_stop_threshold:
                    break
                continue

            empty_streak = 0

            # validation
            if populated < 2:
                for col_name, value in zip(target_fields, values):
                    if not value:
                        issues.append(
                            make_issue(
                                sheet.title,
                                i,
                                "records",
                                col_name,
                                "rejected_row_missing_value",
                                "",
                                "Row was skipped because it had fewer than 2 populated record fields.",
                            )
                        )
                continue

            record_data = dict(zip(target_fields, values))
            for optional_field in OPTIONAL_RECORD_FIELDS:
                if optional_field not in record_data:
                    record_data[optional_field] = ""

            for col_name in target_fields:
                value = record_data.get(col_name, "")
                if not value:
                    issues.append(
                        make_issue(
                            sheet.title,
                            i,
                            "records",
                            col_name,
                            "record_value_missing",
                            "",
                            "Record row was extracted, but this field is empty.",
                        )
                    )

            records.append(
                [file_id, sheet.title]
                + [record_data.get(col, "") for col in BASE_RECORD_FIELDS]
                + [record_data.get(col, "") for col in OPTIONAL_RECORD_FIELDS]
            )

    return records, issues


def sql_ident(name):
    return "[" + str(name).replace("]", "]]") + "]"


def sql_string(value):
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, numbers.Integral):
        return str(int(value))
    if isinstance(value, numbers.Real) and not isinstance(value, numbers.Integral):
        if isinstance(value, float) and pd.isna(value):
            return "NULL"
        return str(value)
    if isinstance(value, float) and pd.isna(value):
        return "NULL"
    text = str(value)
    if text == "":
        return "NULL"
    text = text.replace("'", "''")
    return "N'" + text + "'"


def write_table_schema_sql(f, schema_name, table_name, columns):
    full_name = sql_ident(schema_name) + "." + sql_ident(table_name)
    f.write(f"IF OBJECT_ID(N'{schema_name}.{table_name}', N'U') IS NULL\n")
    f.write("BEGIN\n")
    f.write(f"  CREATE TABLE {full_name} (\n")
    for idx, (col_name, col_type, nullable) in enumerate(columns):
        comma = "," if idx < len(columns) - 1 else ""
        null_sql = "NULL" if nullable else "NOT NULL"
        f.write(f"    {sql_ident(col_name)} {col_type} {null_sql}{comma}\n")
    f.write("  );\n")
    f.write("END\n")
    f.write("GO\n\n")


def write_table_inserts_sql(f, schema_name, table_name, df, chunk_size=200):
    if df is None or df.empty:
        return 0

    cols = list(df.columns)
    full_name = sql_ident(schema_name) + "." + sql_ident(table_name)
    col_list = ", ".join(sql_ident(c) for c in cols)

    total = 0
    batch = []
    for row in df.itertuples(index=False, name=None):
        values = ", ".join(sql_string(v) for v in row)
        batch.append("(" + values + ")")
        total += 1
        if len(batch) >= chunk_size:
            f.write(f"INSERT INTO {full_name} ({col_list}) VALUES\n")
            f.write(",\n".join(batch))
            f.write(";\n")
            batch = []

    if batch:
        f.write(f"INSERT INTO {full_name} ({col_list}) VALUES\n")
        f.write(",\n".join(batch))
        f.write(";\n")

    f.write("GO\n\n")
    return total


def export_offline_sql(sql_dir, meta_df, rec_df, img_df, incomplete_df, schema_name="dbo"):
    os.makedirs(sql_dir, exist_ok=True)
    schema_path = os.path.join(sql_dir, "schema.sql")
    data_path = os.path.join(sql_dir, "data.sql")

    meta_cols = [("FileID", "INT", False)] + [(c, "NVARCHAR(MAX)", True) for c in METADATA_FIELDS]
    rec_cols = [("FileID", "INT", False), ("شیت", "NVARCHAR(255)", True)] + [
        (c, "NVARCHAR(MAX)", True) for c in (BASE_RECORD_FIELDS + OPTIONAL_RECORD_FIELDS)
    ]
    img_cols = [
        ("ImageID", "INT", False),
        ("FileID", "INT", False),
        ("FileName", "NVARCHAR(260)", True),
        ("شیت", "NVARCHAR(255)", True),
        ("ImageName", "NVARCHAR(260)", True),
        ("ImagePath", "NVARCHAR(1024)", True),
        ("EmbeddedPath", "NVARCHAR(1024)", True),
    ]
    incomplete_cols = [
        ("FileID", "INT", False),
        ("FileName", "NVARCHAR(260)", True),
        ("SheetName", "NVARCHAR(255)", True),
        ("ExcelRow", "INT", True),
        ("FieldType", "NVARCHAR(50)", True),
        ("Column", "NVARCHAR(255)", True),
        ("IssueType", "NVARCHAR(100)", True),
        ("CellHint", "NVARCHAR(50)", True),
        ("Details", "NVARCHAR(MAX)", True),
    ]

    with open(schema_path, "w", encoding="utf-8") as f:
        f.write("SET NOCOUNT ON;\n")
        f.write("SET XACT_ABORT ON;\n")
        f.write("GO\n\n")
        write_table_schema_sql(f, schema_name, "FileMetadata", meta_cols)
        write_table_schema_sql(f, schema_name, "FileRecords", rec_cols)
        write_table_schema_sql(f, schema_name, "FileImages", img_cols)
        write_table_schema_sql(f, schema_name, "IncompleteInformation", incomplete_cols)

    if "ExcelRow" in incomplete_df.columns:
        incomplete_df = incomplete_df.copy()
        incomplete_df["ExcelRow"] = pd.to_numeric(incomplete_df["ExcelRow"], errors="coerce")

    with open(data_path, "w", encoding="utf-8") as f:
        f.write("SET NOCOUNT ON;\n")
        f.write("SET XACT_ABORT ON;\n")
        f.write("GO\n\n")
        meta_count = write_table_inserts_sql(f, schema_name, "FileMetadata", meta_df)
        rec_count = write_table_inserts_sql(f, schema_name, "FileRecords", rec_df)
        img_count = write_table_inserts_sql(f, schema_name, "FileImages", img_df)
        incomplete_count = write_table_inserts_sql(f, schema_name, "IncompleteInformation", incomplete_df)

    return {
        "schema_sql": schema_path,
        "data_sql": data_path,
        "meta_rows": meta_count,
        "record_rows": rec_count,
        "image_rows": img_count,
        "incomplete_rows": incomplete_count,
    }


# ==============================
# MAIN
# ==============================
def main():
    args = parse_args()

    input_folder = args.input
    output_folder = args.output

    os.makedirs(output_folder, exist_ok=True)

    logging.basicConfig(
        filename=os.path.join(output_folder, "process.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )

    files = glob.glob(os.path.join(input_folder, "*.xlsx"))
    total_files = len(files)

    if total_files == 0:
        print(f"No .xlsx files found in: {input_folder}")
        return

    metadata_rows = []
    record_rows = []
    image_rows = []
    incomplete_rows = []
    success_count = 0
    failed_count = 0
    next_image_id = 1

    for i, file in enumerate(files, 1):
        try:
            print(f"Processing {i}/{total_files}: {file}")

            wb = load_workbook(file, data_only=True, read_only=True)

            file_id = i

            metadata, metadata_issues = extract_metadata(wb)
            metadata_rows.append([file_id] + [metadata[k] for k in METADATA_FIELDS])

            records, record_issues = extract_records(wb, file_id)
            record_rows.extend(records)

            file_images, next_image_id = extract_images_from_excel(file, file_id, output_folder, next_image_id)
            image_rows.extend(file_images)

            file_name = os.path.basename(file)
            for issue in metadata_issues + record_issues:
                issue_row = {"FileID": file_id, "FileName": file_name}
                issue_row.update(issue)
                incomplete_rows.append(issue_row)

            wb.close()
            success_count += 1

        except Exception as e:
            failed_count += 1
            logging.error(f"Error in {file}: {str(e)}")

    # SAVE
    meta_df = pd.DataFrame(metadata_rows, columns=["FileID"] + METADATA_FIELDS)
    rec_df = pd.DataFrame(record_rows, columns=RECORD_OUTPUT_COLUMNS)
    img_df = pd.DataFrame(image_rows, columns=IMAGE_OUTPUT_COLUMNS)
    incomplete_df = pd.DataFrame(incomplete_rows, columns=INCOMPLETE_COLUMNS)

    metadata_xlsx = os.path.join(output_folder, "file_metadata.xlsx")
    records_xlsx = os.path.join(output_folder, "file_records.xlsx")
    images_xlsx = os.path.join(output_folder, "file_images.xlsx")
    incomplete_xlsx = os.path.join(output_folder, "incomplete_information.xlsx")
    metadata_csv = os.path.join(output_folder, "file_metadata.csv")
    records_csv = os.path.join(output_folder, "file_records.csv")
    images_csv = os.path.join(output_folder, "file_images.csv")
    incomplete_csv = os.path.join(output_folder, "incomplete_information.csv")
    log_file = os.path.join(output_folder, "process.log")

    meta_df.to_excel(metadata_xlsx, index=False)
    rec_df.to_excel(records_xlsx, index=False)
    img_df.to_excel(images_xlsx, index=False)
    incomplete_df.to_excel(incomplete_xlsx, index=False)

    meta_df.to_csv(metadata_csv, index=False, encoding="utf-8-sig")
    rec_df.to_csv(records_csv, index=False, encoding="utf-8-sig")
    img_df.to_csv(images_csv, index=False, encoding="utf-8-sig")
    incomplete_df.to_csv(incomplete_csv, index=False, encoding="utf-8-sig")

    sql_export = None
    if args.export_sql:
        sql_dir = args.export_sql.strip()
        sql_export = export_offline_sql(sql_dir, meta_df, rec_df, img_df, incomplete_df)

    print("\nFinal Report")
    print(f"Total files:   {total_files}")
    print(f"Success count: {success_count}")
    print(f"Failed count:  {failed_count}")
    print(f"Metadata rows: {len(metadata_rows)}")
    print(f"Record rows:   {len(record_rows)}")
    print(f"Image rows:    {len(image_rows)}")
    print(f"Incomplete issues: {len(incomplete_rows)}")
    print("\nOutput files:")
    print(f"- {metadata_xlsx}")
    print(f"- {records_xlsx}")
    print(f"- {images_xlsx}")
    print(f"- {incomplete_xlsx}")
    print(f"- {metadata_csv}")
    print(f"- {records_csv}")
    print(f"- {images_csv}")
    print(f"- {incomplete_csv}")
    print(f"- {log_file}")
    if sql_export:
        print("\nOffline SQL export:")
        print(f"- {sql_export['schema_sql']}")
        print(f"- {sql_export['data_sql']}")


if __name__ == "__main__":
    main()
