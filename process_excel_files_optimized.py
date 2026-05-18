#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
process_excel_files.py

Batch-extract metadata + records from many .xlsx files into two flat tables
ready for SQL Server ingestion (XLSX + CSV outputs).

Usage:
    python process_excel_files.py --input "C:\excel_files" --output "C:\excel_output"
"""

import os
import re
import glob
import argparse
import logging
from datetime import datetime
from typing import Dict, List, Any

import pandas as pd
from openpyxl import load_workbook


# ==============================
# ARGUMENTS
# ==============================
def parse_args():
    parser = argparse.ArgumentParser(description="Process Excel files into structured datasets.")
    parser.add_argument("--input", required=True, help="Input folder containing .xlsx files")
    parser.add_argument("--output", required=True, help="Output folder for results")
    return parser.parse_args()


# ==============================
# CONSTANTS
# ==============================
METADATA_FIELDS = [
    "تاریخ انتشار",
    "شماره صفحه",
    "کد پروژه",
    "نام سازنده",
    "شماره بازنگری",
    "تهیه کننده",
    "تایید کننده",
    "نام قطعه",
    "شماره فنی",
    "نام خودرو",
]

RECORD_FIELDS = [
    "مشخصات کنترلی",
    "شرح کنترل",
    "درجه اهمیت",
    "استاندارد مرجع",
    "محدوده قابل قبول",
    "روش نمونه گیری (تناوب - تعداد) توضیحات",
]


# ==============================
# UTILS
# ==============================
def normalize(text):
    if text is None:
        return ""
    return str(text).strip().replace("ي", "ی").replace("ك", "ک")


def extract_part_number(text):
    text = normalize(text).upper()
    match = re.search(r"\b[0-9A-Z]{6,}(?:-[0-9A-Z]{2,})+\b", text)
    return match.group(0) if match else ""


# ==============================
# METADATA EXTRACTION
# ==============================
def extract_metadata(wb):
    result = {k: "" for k in METADATA_FIELDS}

    patterns = {
        "تاریخ انتشار": [("تاریخ", "انتشار"), ("انتشار",), ("تاریخ",)],
        "شماره صفحه": [("شماره", "صفحه")],
        "کد پروژه": [("کد", "پروژه")],
        "نام سازنده": [("نام", "سازنده")],
        "شماره بازنگری": [("شماره", "بازنگری")],
        "تهیه کننده": [("تهیه", "کننده")],
        "تایید کننده": [("تایید", "کننده")],
        "نام قطعه": [("نام", "قطعه")],
        "شماره فنی": [("شماره", "فنی")],
        "نام خودرو": [("نام", "خودرو")],
    }

    for sheet in wb.worksheets[2:]:
        for row in sheet.iter_rows(max_row=200, max_col=50):
            for cell in row:
                text = normalize(cell.value)

                for field, pats in patterns.items():
                    if result[field]:
                        continue

                    for pat in pats:
                        if all(p in text for p in pat):
                            val = ""

                            # same cell
                            if ":" in text:
                                val = text.split(":")[-1].strip()

                            # next cell
                            if not val:
                                try:
                                    val = normalize(row[cell.column].value)
                                except Exception:
                                    val = ""

                            if val:
                                if field == "شماره فنی":
                                    val = extract_part_number(val)

                                result[field] = val
                                break

    return result


# ==============================
# RECORDS EXTRACTION
# ==============================
def detect_header(row):
    row_text = [normalize(x) for x in row if x]
    score = sum(1 for col in RECORD_FIELDS if any(col[:4] in cell for cell in row_text))
    return score >= 3


def extract_records(wb, file_id):
    records = []

    for sheet in wb.worksheets[2:]:
        header = None
        header_idx = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if detect_header(row):
                header = row
                header_idx = i
                break

        if not header:
            continue

        col_map = {}
        for idx, col in enumerate(header):
            col_text = normalize(col)
            for target in RECORD_FIELDS:
                if target[:4] in col_text:
                    col_map[target] = idx

        for i, row in enumerate(sheet.iter_rows(min_row=header_idx + 2, values_only=True)):
            values = []

            for col in RECORD_FIELDS:
                idx = col_map.get(col)
                val = normalize(row[idx]) if idx is not None and idx < len(row) else ""
                values.append(val)

            # validation
            if sum(1 for v in values if v) < 2:
                continue

            records.append([file_id] + values)

    return records


# ==============================
# MAIN
# ==============================
def main():
    args = parse_args()

    input_folder = args.input
    output_folder = args.output

    os.makedirs(output_folder, exist_ok=True)

    logging.basicConfig(
        filename=os.path.join(output_folder, "process.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )

    files = glob.glob(os.path.join(input_folder, "*.xlsx"))

    metadata_rows = []
    record_rows = []

    for i, file in enumerate(files, 1):
        try:
            print(f"Processing {i}/{len(files)}: {file}")

            wb = load_workbook(file, data_only=True)

            file_id = i

            metadata = extract_metadata(wb)
            metadata_rows.append([file_id] + [metadata[k] for k in METADATA_FIELDS])

            records = extract_records(wb, file_id)
            record_rows.extend(records)

        except Exception as e:
            logging.error(f"Error in {file}: {str(e)}")

    # SAVE
    meta_df = pd.DataFrame(metadata_rows, columns=["FileID"] + METADATA_FIELDS)
    rec_df = pd.DataFrame(record_rows, columns=["FileID"] + RECORD_FIELDS)

    meta_df.to_excel(os.path.join(output_folder, "file_metadata.xlsx"), index=False)
    rec_df.to_excel(os.path.join(output_folder, "file_records.xlsx"), index=False)

    meta_df.to_csv(os.path.join(output_folder, "file_metadata.csv"), index=False, encoding="utf-8-sig")
    rec_df.to_csv(os.path.join(output_folder, "file_records.csv"), index=False, encoding="utf-8-sig")

    print("DONE ✅")


if __name__ == "__main__":
    main()
